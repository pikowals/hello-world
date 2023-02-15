import openpyxl
sheetData = []
wb = openpyxl.load_workbook('Items.xlsx')
sheet =wb.active
x=0
for i in range(1,sheet.max_column+1):
    sheetData.append([])
    for j in range(1,sheet.max_row+1):
        sheetData[x].append(sheet.cell(row=j, column=i).value)
        sheet.cell(row=j, column=i).value = None
    x += 1
for i in range(len(sheetData)):
    for j in range(len(sheetData[i])):
        sheet.cell(row=i+1,column =j+1).value = sheetData[i][j]
wb.save('calculator_updated.xlsx')
print(sheetData)