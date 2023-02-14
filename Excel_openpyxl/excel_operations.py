
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
wb = openpyxl.load_workbook('example.xlsx')
sheet_obj = wb.active
sheets = wb.sheetnames
print(sheets)
print(sheet_obj)
cell_obj = sheet_obj.cell(row=1,column=1)
print(cell_obj.value)
print(sheet_obj.max_row)        # metoda zwracajaca zmienna przechowujaca ilość wierszy
print(sheet_obj.max_column)     # metoda zwracajaca zmienna przechowujaca ilość kolumn

print(tuple(sheet_obj['A1':'C3']))
for rowOfCellObject in sheet_obj['A1':'C3']:
    for cellObj in rowOfCellObject:
        print(cellObj.coordinate,cellObj.value)
    print('---Koniec Wiersza---')

for i in sheet_obj.columns:
    cur_col = list(i)
    for j in cur_col:
        print(j.value)










# for j in range(1, sheet_obj.max_row+1):
#     cell = sheet_obj.cell(row=j,column=3)
#     #print(cell.value)       # Przechowuje wartośc komórki
#     print(cell.coordinate,cell.value)  # Przechowuje koordynaty danej komórki np. A5
# cell_1 = sheet_obj.cell(row=1,column=1)
# print(type(cell_1.value))   # Daty są przez moduł openpyxl automatycznie interpretowane jako typ datetime
# for i in range(1,sheet_obj.max_column+1):
#     cell = sheet_obj.cell(row=1,column =i)
#     print(cell.value)
#
#
# print(get_column_letter(sheet_obj.max_column)) # Konwersja liczb na litery w tym przypadku kolumny
# print(column_index_from_string('AAH'))           # Konwersja liter na liczby w tym przypadku kolumny
# print(column_index_from_string(get_column_letter(sheet_obj.max_column))) # To samo co powyżej
