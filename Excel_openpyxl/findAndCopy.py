"""Program is finding values from reqular expression and copy
it to another file"""

import openpyxl, re
productRegex = re.compile(r'''((^G).+)''', re.DOTALL)
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']
sheet1 = wb.create_sheet('Sheet1')
num = 1
for rowNr in range(2,sheet.max_row +1):
    productName = sheet.cell(row=rowNr,column=1).value
    if productRegex.search(productName):
        sheet1.cell(row=num,column=1).value = productName
        num +=1
wb.save('Copied_items.xlsx')
print(num)