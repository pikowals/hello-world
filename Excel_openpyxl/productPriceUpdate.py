"""Program aktualizuje ceny produktów w excelu w zależnośći danych zawartych
w słowniku i zapisuje ilość wykonanych zmian"""

import openpyxl
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

PRICE_UPDATES = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27} # Elegancko zapisane zmiany cen w słowniku
change = 0                                                              # W przypadku dodatkowych zmian edytujemy tylko tą cześć
for rowNr in range(2,sheet.max_row+1):                          # Iterujemy przez wszystkie wiersze
    productName = sheet.cell(row= rowNr,column = 1).value       # Zmienna productName przechowuje wartość pobrana z komórki
    if productName in PRICE_UPDATES:                            # Jeśli trafimy na wartość odpowiadajacą kluczowi słownika
        sheet[f'B{rowNr}'] = PRICE_UPDATES[productName]         # Aktualizujemy wartość tego wiersza ale w kolumnie B
        change = change +1
wb.save('updated_produce_sales.xlsx')                           # Zapisujemy plik
print(change)                                                   # Ilość zmian dokonanych w excelu
