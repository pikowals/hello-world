
from openpyxl.styles import Font, NamedStyle
# wb = openpyxl.Workbook()
# print(wb.sheetnames)
# sheet = wb.active
# sheet.title = 'Spam'
# print(sheet.title)
# wb.create_sheet('Sheet1')
# print(wb.sheetnames)
# wb.create_sheet(index=0,title='Pierwszy arkusz') # Tworzy arkusz pod odpowiednim indeksem
# print(wb.sheetnames)
# wb.create_sheet(index=2,title='Drugi arkusz')    # Tworzy arkusz pod odpowiednim indeksem
# print(wb.sheetnames)
# del wb['Drugi arkusz']
# print(wb.sheetnames)
# del wb['Spam']
# print(wb.sheetnames)
# sheet = wb['Sheet1']
# sheet['A1'] = 'Witaj Świecie-nowość'
# print(sheet['A1'].value)
# wb.save('Example_copy.xlsx')

# wb = openpyxl.Workbook()
# sheet = wb['Sheet']
# italic24font = Font(size=24, italic=True)
# styleObj = NamedStyle(font=italic24font)
# print(type(styleObj))
# sheet['A'].font = styleObj
# sheet['A1'] = 'Witaj Świecie'
# wb.save('styled.xlsx')
from openpyxl.utils import get_column_letter, column_index_from_string
# import openpyxl
# wb = openpyxl.Workbook()
# sheet = wb.active
# print('Podaj wielkość N a wygeneruje tabliczke mnożenia w excelu o wielkośći NxN... ')
# N = input()
# for i in range(1,int(N)+1):
#     sheet.cell(row=i+1,column=1).value = i*1
#     for j in range(1,int(N)+1):
#         sheet.cell(row=i+1,column=j+1).value = i*j
#         print(str(i*j) + ' ', end ='')
#     print('\n')
# wb.save('calulator.xlsx')

import openpyxl
sheetData = []
wb = openpyxl.load_workbook('Items.xlsx')
sheet =wb.active
x=0
for i in range(1,sheet.max_column+1):
    sheetData.append([])
    for j in range(1,sheet.max_row+1):
        sheetData[x].append(sheet.cell(row=j, column=i).value)
        sheet.cell(row=j, column=i).value = None
    x += 1
for i in range(len(sheetData)):
    for j in range(len(sheetData[i])):
        sheet.cell(row=i+1,column =j+1).value = sheetData[i][j]
wb.save('calculator_updated.xlsx')
print(sheetData)