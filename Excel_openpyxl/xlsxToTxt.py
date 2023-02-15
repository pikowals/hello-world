import openpyxl

wb = openpyxl.load_workbook('txt_to_xlsx.xlsx')
sheet = wb.active

for i in range(1, sheet.max_column+1):
    file = open(f'copied_from_excel_{i}.txt','w',encoding="utf8")
    for j in range(1, sheet.max_row+1):
        if sheet.cell(row=j, column=i).value == None:
            continue
        file.write(sheet.cell(row=j, column=i).value)
    file.close()

